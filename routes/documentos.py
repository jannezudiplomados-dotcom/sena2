import os
import io
import zipfile
import shutil
import subprocess
from datetime import datetime
from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, send_file, session)
from werkzeug.utils import secure_filename
from routes.auth import login_required
from models import (obtener_plantillas, obtener_plantilla, crear_plantilla,
                    eliminar_plantilla as del_plantilla, obtener_fichas,
                    obtener_programas, obtener_usuarios_por_ficha,
                    obtener_usuarios_por_programa, obtener_usuario, registrar_actividad)
from config import PLANTILLAS_FOLDER, GENERADOS_FOLDER, FIRMAS_FOLDER

documentos_bp = Blueprint('documentos', __name__, url_prefix='/documentos')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ['docx', 'xlsx', 'xls']


def _nombre_seguro(usuario):
    """Construye un nombre de archivo seguro (evita inyeccion de rutas/comandos)."""
    base = f"{usuario['identificacion']}_{usuario['apellidos']}_{usuario['nombre']}"
    return secure_filename(base) or 'documento'


def generar_docx(plantilla_path, contexto, output_path):
    """Genera un documento DOCX a partir de una plantilla y contexto."""
    from docxtpl import DocxTemplate, InlineImage
    from docx.shared import Mm

    doc = DocxTemplate(plantilla_path)

    if contexto.get('firma_ruta'):
        try:
            contexto['firma'] = InlineImage(doc, contexto['firma_ruta'], width=Mm(40))
        except Exception as e:
            print(f"No se pudo cargar la imagen de la firma: {e}")
            contexto['firma'] = ''
    else:
        contexto['firma'] = ''

    doc.render(contexto)
    doc.save(output_path)
    return output_path


def generar_xlsx(plantilla_path, contexto, output_path):
    """Genera un documento Excel a partir de una plantilla y contexto."""
    import openpyxl
    from jinja2 import Template

    wb = openpyxl.load_workbook(plantilla_path)

    if contexto.get('firma_ruta'):
        contexto['firma'] = '[[FIRMA_AQUI]]'
    else:
        contexto['firma'] = ''

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                    try:
                        template = Template(cell.value)
                        cell.value = template.render(**contexto)
                    except Exception as e:
                        print(f"Error al renderizar celda {cell.coordinate}: {e}")

    if contexto.get('firma_ruta') and os.path.exists(contexto['firma_ruta']):
        from openpyxl.drawing.image import Image as OpenpyxlImage
        try:
            firma_img = OpenpyxlImage(contexto['firma_ruta'])
            firma_img.width = 150
            firma_img.height = 75
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and '[[FIRMA_AQUI]]' in cell.value:
                            cell.value = cell.value.replace('[[FIRMA_AQUI]]', '')
                            sheet.add_image(firma_img, cell.coordinate)
        except Exception as e:
            print(f"Error cargando imagen de firma para excel: {e}")

    wb.save(output_path)
    return output_path


def _soffice_a_pdf(origen_abs, pdf_path_abs):
    """Convierte con LibreOffice headless (multiplataforma, sin ejecutar codigo dinamico)."""
    soffice = shutil.which('soffice') or shutil.which('libreoffice')
    if not soffice:
        return False
    output_dir = os.path.dirname(pdf_path_abs)
    try:
        subprocess.run(
            [soffice, '--headless', '--convert-to', 'pdf', '--outdir', output_dir, origen_abs],
            capture_output=True, text=True, timeout=120, check=False
        )
    except Exception as e:
        print(f"LibreOffice fallo: {e}")
        return False
    base_name = os.path.splitext(os.path.basename(origen_abs))[0]
    generado = os.path.join(output_dir, f"{base_name}.pdf")
    if os.path.exists(generado) and generado != pdf_path_abs:
        shutil.move(generado, pdf_path_abs)
    return os.path.exists(pdf_path_abs)


def docx_a_pdf(docx_path, pdf_path):
    """DOCX -> PDF. 1) docx2pdf (Word en Windows) 2) LibreOffice 3) reportlab.
    NOTA: ya no se ejecuta codigo dinamico con 'python -c' (evita inyeccion)."""
    docx_path_abs = os.path.abspath(docx_path)
    pdf_path_abs = os.path.abspath(pdf_path)

    # Metodo 1: docx2pdf (usa Word en Windows; libreria, sin script dinamico)
    try:
        from docx2pdf import convert
        convert(docx_path_abs, pdf_path_abs)
        if os.path.exists(pdf_path_abs):
            return True
    except Exception as e:
        print(f"docx2pdf no disponible o fallo: {e}")

    # Metodo 2: LibreOffice headless
    try:
        if _soffice_a_pdf(docx_path_abs, pdf_path_abs):
            return True
    except Exception as e:
        print(f"LibreOffice fallo: {e}")

    # Metodo 3: Fallback reportlab (calidad basica)
    try:
        return _docx_a_pdf_fallback(docx_path_abs, pdf_path_abs)
    except Exception as e:
        print(f"Error en fallback docx->pdf: {e}")
        return False


def _docx_a_pdf_fallback(docx_path, pdf_path):
    """Fallback: python-docx (leer) + reportlab (escribir). No preserva formato avanzado."""
    from docx import Document
    from docx.oxml.ns import qn
    from docx.table import Table as DocxTable
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_JUSTIFY

    doc_word = Document(docx_path)
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle('BodyCustom', parent=styles['Normal'],
                                fontSize=10, leading=14, alignment=TA_JUSTIFY, spaceAfter=6)
    heading_style = ParagraphStyle('HeadingCustom', parent=styles['Heading2'],
                                   fontSize=14, leading=18, spaceAfter=10, spaceBefore=12,
                                   textColor=colors.HexColor('#39A900'))
    cell_style = ParagraphStyle('CellCustom', parent=styles['Normal'],
                                fontSize=9, leading=11, alignment=TA_LEFT)

    pdf_doc = SimpleDocTemplate(pdf_path, pagesize=letter,
                                leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                                topMargin=0.75 * inch, bottomMargin=0.75 * inch)
    elements = []

    def esc(t):
        return t.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    for element in doc_word.element.body:
        tag = element.tag.split('}')[-1] if '}' in element.tag else element.tag
        if tag == 'p':
            style_elem = element.find(qn('w:pPr'))
            is_heading = False
            if style_elem is not None:
                style_name = style_elem.find(qn('w:pStyle'))
                if style_name is not None:
                    val = style_name.get(qn('w:val'), '')
                    if 'Heading' in val or 'heading' in val or 'Titulo' in val:
                        is_heading = True
            text = ''.join(node.text or '' for node in element.iter()
                           if node.text and node.tag.endswith('}t')).strip()
            if text:
                elements.append(Paragraph(esc(text), heading_style if is_heading else body_style))
        elif tag == 'tbl':
            try:
                tbl = DocxTable(element, doc_word)
                data = []
                for row in tbl.rows:
                    data.append([Paragraph(esc(cell.text.strip()), cell_style) for cell in row.cells])
                if data:
                    num_cols = max(len(r) for r in data)
                    available = letter[0] - 1.5 * inch
                    col_w = available / max(num_cols, 1)
                    table = Table(data, colWidths=[col_w] * num_cols)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39A900')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ]))
                    elements.append(Spacer(1, 6))
                    elements.append(table)
                    elements.append(Spacer(1, 6))
            except Exception as e:
                print(f"Error procesando tabla en DOCX: {e}")

    if not elements:
        elements.append(Paragraph('Documento generado desde plantilla DOCX', body_style))

    pdf_doc.build(elements)
    return os.path.exists(pdf_path)


def xlsx_a_pdf(xlsx_path, pdf_path):
    """XLSX -> PDF. 1) LibreOffice 2) reportlab. Sin ejecutar codigo dinamico."""
    xlsx_path_abs = os.path.abspath(xlsx_path)
    pdf_path_abs = os.path.abspath(pdf_path)
    try:
        if _soffice_a_pdf(xlsx_path_abs, pdf_path_abs):
            return True
    except Exception as e:
        print(f"LibreOffice fallo: {e}")
    try:
        return _xlsx_a_pdf_fallback(xlsx_path_abs, pdf_path_abs)
    except Exception as e:
        print(f"Error en fallback xlsx->pdf: {e}")
        return False


def _xlsx_a_pdf_fallback(xlsx_path, pdf_path):
    """Fallback: openpyxl (leer) + reportlab (escribir). Divide tablas grandes en bloques."""
    import openpyxl
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=6, leading=8, alignment=TA_LEFT)

    doc = SimpleDocTemplate(pdf_path, pagesize=landscape(letter),
                            leftMargin=0.3 * inch, rightMargin=0.3 * inch,
                            topMargin=0.4 * inch, bottomMargin=0.4 * inch)
    elements = []
    MAX_COLS, MAX_CELL_LEN, ROWS_PER_CHUNK = 15, 60, 40

    def esc(t):
        return t.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    for sheet in wb.worksheets:
        elements.append(Paragraph(f"Hoja: {sheet.title}", title_style))
        elements.append(Spacer(1, 6))
        all_rows = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                        min_col=1, max_col=sheet.max_column))
        if not all_rows:
            continue
        real_max_col = 0
        for row in all_rows[:50]:
            for idx, cell in enumerate(row):
                if cell.value is not None and str(cell.value).strip():
                    real_max_col = max(real_max_col, idx + 1)
        num_cols = min(real_max_col, MAX_COLS) if real_max_col > 0 else min(len(all_rows[0]), MAX_COLS)
        if num_cols == 0:
            continue
        data = []
        for row in all_rows:
            row_data = []
            for idx, cell in enumerate(row):
                if idx >= num_cols:
                    break
                text = str(cell.value if cell.value is not None else '').strip()
                if len(text) > MAX_CELL_LEN:
                    text = text[:MAX_CELL_LEN] + '...'
                row_data.append(Paragraph(esc(text), cell_style))
            data.append(row_data)
        for row_data in data:
            while len(row_data) < num_cols:
                row_data.append(Paragraph('', cell_style))
        available_width = landscape(letter)[0] - 0.6 * inch
        col_widths = [available_width / max(num_cols, 1)] * num_cols
        header_row = data[0] if data else []
        body_rows = data[1:] if len(data) > 1 else []
        for chunk_start in range(0, max(len(body_rows), 1), ROWS_PER_CHUNK):
            chunk = body_rows[chunk_start:chunk_start + ROWS_PER_CHUNK]
            table_data = [header_row] + chunk if chunk else [header_row]
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39A900')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 6),
                ('FONTSIZE', (0, 1), (-1, -1), 5),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 8))

    if elements:
        doc.build(elements)
        return os.path.exists(pdf_path)
    return False


def _contexto_usuario(u):
    contexto = {
        'tipo_documento': u.get('tipo_documento', 'CC'),
        'nombre': u['nombre'],
        'apellidos': u['apellidos'],
        'identificacion': u['identificacion'],
        'telefono': u.get('telefono', ''),
        'correo_institucional': u.get('correo_institucional', ''),
        'correo_personal': u.get('correo_personal', ''),
        'correo_electronico': u.get('correo_institucional', '') or u.get('correo_personal', ''),
        'direccion_residencia': u.get('direccion_residencia', ''),
        'numero_ficha': u.get('numero_ficha', ''),
        'programa_formacion': u.get('nombre_programa', ''),
        'ficha': u.get('numero_ficha', ''),
        'programa': u.get('nombre_programa', ''),
        'fecha': datetime.now().strftime('%d/%m/%Y'),
    }
    if u.get('firma_imagen'):
        firma_path = os.path.join(FIRMAS_FOLDER, secure_filename(u['firma_imagen']))
        if os.path.exists(firma_path):
            contexto['firma_ruta'] = firma_path
    return contexto


@documentos_bp.route('/')
@login_required
def index():
    plantillas = obtener_plantillas()
    fichas = obtener_fichas()
    programas = obtener_programas()
    return render_template('documentos/plantillas.html',
                           plantillas=plantillas, fichas=fichas, programas=programas)


@documentos_bp.route('/subir_plantilla', methods=['POST'])
@login_required
def subir_plantilla():
    if 'archivo' not in request.files:
        flash('No se selecciono ningun archivo.', 'danger')
        return redirect(url_for('documentos.index'))

    file = request.files['archivo']
    if file.filename == '':
        flash('No se selecciono ningun archivo.', 'danger')
        return redirect(url_for('documentos.index'))

    if file and allowed_file(file.filename):
        nombre = request.form.get('nombre', '').strip() or file.filename
        descripcion = request.form.get('descripcion', '').strip()
        filename = secure_filename(file.filename)
        base, ext = os.path.splitext(filename)
        filename = f"{base}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
        filepath = os.path.join(PLANTILLAS_FOLDER, filename)
        file.save(filepath)
        try:
            nuevo_id = crear_plantilla(nombre, filename, descripcion)
            registrar_actividad(session.get('admin'), 'crear', 'plantilla', nuevo_id, nombre, request.remote_addr)
            flash('Plantilla subida exitosamente.', 'success')
        except Exception:
            flash('Error al guardar la plantilla.', 'danger')
    else:
        flash('Solo se permiten archivos .docx o Excel (.xlsx, .xls)', 'danger')

    return redirect(url_for('documentos.index'))


@documentos_bp.route('/eliminar_plantilla/<int:id>', methods=['POST'])
@login_required
def eliminar_plantilla(id):
    plantilla = obtener_plantilla(id)
    if plantilla:
        filepath = os.path.join(PLANTILLAS_FOLDER, secure_filename(plantilla['archivo']))
        if os.path.exists(filepath):
            os.remove(filepath)
        try:
            del_plantilla(id)
            registrar_actividad(session.get('admin'), 'eliminar', 'plantilla', id, None, request.remote_addr)
            flash('Plantilla eliminada.', 'success')
        except Exception:
            flash('Error al eliminar la plantilla.', 'danger')

    return redirect(url_for('documentos.index'))


@documentos_bp.route('/generar', methods=['POST'])
@login_required
def generar():
    plantilla_id = request.form.get('plantilla_id', '')
    tipo = request.form.get('tipo', '')
    grupo_id = request.form.get('grupo_id', '')
    formato = request.form.get('formato', 'docx')

    if not plantilla_id or not tipo or not grupo_id:
        flash('Debe seleccionar plantilla, tipo y grupo.', 'danger')
        return redirect(url_for('documentos.index'))

    plantilla = obtener_plantilla(int(plantilla_id))
    if not plantilla:
        flash('Plantilla no encontrada.', 'danger')
        return redirect(url_for('documentos.index'))

    plantilla_path = os.path.join(PLANTILLAS_FOLDER, secure_filename(plantilla['archivo']))
    if not os.path.exists(plantilla_path):
        flash('El archivo de plantilla no existe.', 'danger')
        return redirect(url_for('documentos.index'))

    if tipo == 'ficha':
        usuarios = obtener_usuarios_por_ficha(int(grupo_id))
    elif tipo == 'programa':
        usuarios = obtener_usuarios_por_programa(int(grupo_id))
    else:
        flash('Tipo de generacion no valido.', 'danger')
        return redirect(url_for('documentos.index'))

    if not usuarios:
        flash('No se encontraron aprendices para generar documentos.', 'warning')
        return redirect(url_for('documentos.index'))

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    carpeta_gen = os.path.join(GENERADOS_FOLDER, f"gen_{timestamp}")
    os.makedirs(carpeta_gen, exist_ok=True)

    archivos_generados = []
    ext = os.path.splitext(plantilla_path)[1].lower()

    for u in usuarios:
        contexto = _contexto_usuario(u)
        nombre_doc = _nombre_seguro(u)
        try:
            if ext in ['.xlsx', '.xls']:
                xlsx_path = os.path.join(carpeta_gen, f"{nombre_doc}{ext}")
                generar_xlsx(plantilla_path, contexto, xlsx_path)
                if formato == 'pdf':
                    pdf_path = os.path.join(carpeta_gen, f"{nombre_doc}.pdf")
                    if xlsx_a_pdf(xlsx_path, pdf_path):
                        archivos_generados.append(pdf_path)
                        os.remove(xlsx_path)
                    else:
                        archivos_generados.append(xlsx_path)
                        flash(f'No se pudo convertir a PDF para {u["nombre"]}. Se genero Excel.', 'warning')
                else:
                    archivos_generados.append(xlsx_path)
            else:
                docx_path = os.path.join(carpeta_gen, f"{nombre_doc}.docx")
                generar_docx(plantilla_path, contexto, docx_path)
                if formato == 'pdf':
                    pdf_path = os.path.join(carpeta_gen, f"{nombre_doc}.pdf")
                    if docx_a_pdf(docx_path, pdf_path):
                        archivos_generados.append(pdf_path)
                        os.remove(docx_path)
                    else:
                        archivos_generados.append(docx_path)
                        flash(f'No se pudo convertir a PDF para {u["nombre"]}. Se genero DOCX.', 'warning')
                else:
                    archivos_generados.append(docx_path)
        except Exception as e:
            print(f"Error generando documento para {u.get('identificacion')}: {e}")
            flash(f'Error generando documento para {u["nombre"]}.', 'danger')

    if not archivos_generados:
        flash('No se generaron documentos.', 'danger')
        return redirect(url_for('documentos.index'))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for archivo in archivos_generados:
            zipf.write(archivo, os.path.basename(archivo))
    zip_buffer.seek(0)

    zip_nombre = f"documentos_{secure_filename(tipo)}_{secure_filename(str(grupo_id))}_{timestamp}.zip"
    zip_path = os.path.join(GENERADOS_FOLDER, zip_nombre)
    with open(zip_path, 'wb') as f:
        f.write(zip_buffer.getvalue())

    registrar_actividad(session.get('admin'), 'generar', tipo, int(grupo_id),
                        f"{len(archivos_generados)} documentos", request.remote_addr)
    flash(f'Se generaron {len(archivos_generados)} documentos exitosamente.', 'success')
    return send_file(zip_path, mimetype='application/zip', as_attachment=True, download_name=zip_nombre)


@documentos_bp.route('/generar_individual/<int:usuario_id>/<int:plantilla_id>')
@login_required
def generar_individual(usuario_id, plantilla_id):
    """Genera un documento individual. Soporta ?formato=pdf."""
    formato = request.args.get('formato', 'original')

    usuario = obtener_usuario(usuario_id)
    plantilla = obtener_plantilla(plantilla_id)
    if not usuario or not plantilla:
        flash('Usuario o plantilla no encontrados.', 'danger')
        return redirect(url_for('documentos.index'))

    plantilla_path = os.path.join(PLANTILLAS_FOLDER, secure_filename(plantilla['archivo']))
    if not os.path.exists(plantilla_path):
        flash('Archivo de plantilla no encontrado.', 'danger')
        return redirect(url_for('documentos.index'))

    contexto = _contexto_usuario(usuario)
    nombre_doc = _nombre_seguro(usuario)
    ext = os.path.splitext(plantilla_path)[1].lower()

    try:
        if ext in ['.xlsx', '.xls']:
            xlsx_path = os.path.join(GENERADOS_FOLDER, f"{nombre_doc}{ext}")
            generar_xlsx(plantilla_path, contexto, xlsx_path)
            if formato == 'pdf':
                pdf_path = os.path.join(GENERADOS_FOLDER, f"{nombre_doc}.pdf")
                if xlsx_a_pdf(xlsx_path, pdf_path):
                    os.remove(xlsx_path)
                    return send_file(pdf_path, as_attachment=True, download_name=f"{nombre_doc}.pdf")
                flash('No se pudo convertir a PDF. Se descarga el Excel.', 'warning')
            return send_file(xlsx_path, as_attachment=True, download_name=f"{nombre_doc}{ext}")
        else:
            docx_path = os.path.join(GENERADOS_FOLDER, f"{nombre_doc}.docx")
            generar_docx(plantilla_path, contexto, docx_path)
            if formato == 'pdf':
                pdf_path = os.path.join(GENERADOS_FOLDER, f"{nombre_doc}.pdf")
                if docx_a_pdf(docx_path, pdf_path):
                    os.remove(docx_path)
                    return send_file(pdf_path, as_attachment=True, download_name=f"{nombre_doc}.pdf")
                flash('No se pudo convertir a PDF. Se descarga el DOCX.', 'warning')
            return send_file(docx_path, as_attachment=True, download_name=f"{nombre_doc}.docx")
    except Exception as e:
        print(f"Error al generar documento individual: {e}")
        flash('Error al generar el documento.', 'danger')
        return redirect(url_for('documentos.index'))
